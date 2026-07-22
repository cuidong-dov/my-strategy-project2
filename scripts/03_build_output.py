#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股板块评估系统 · 小波段输出生成器
读取策略引擎输出的 signal CSV + stats/params JSON，生成 Excel 和 MD。
"""
import pandas as pd, numpy as np, json, os, sys, statistics as st, importlib.util

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
OUT_DIR  = os.path.join(BASE_DIR, "output")
SECTORS  = ["半导体", "消费电子", "通信", "电池", "电力", "煤炭"]

# 导入策略引擎的回测和理由函数
_engine = os.path.join(BASE_DIR, "scripts", "02_swing_engine.py")
_spec = importlib.util.spec_from_file_location("engine", _engine)
_mod = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(_mod)
backtest, reason = _mod.backtest, _mod.reason

def nv(v):
    return "—" if v is None else v

def avg(dd, m):
    vals = [dd[s][m] for s in SECTORS if dd[s][m] is not None]
    return round(st.mean(vals), 1) if vals else "—"

def build():
    stats = json.load(open(f"{OUT_DIR}/swing_stats.json", encoding="utf-8"))
    bp    = json.load(open(f"{OUT_DIR}/swing_params.json", encoding="utf-8"))
    params = bp["params"]

    # 加载信号CSV
    frames = {}
    for s in SECTORS:
        d = pd.read_csv(f"{OUT_DIR}/swing_{s}_signal.csv", parse_dates=["date"])
        d["signal"] = d["signal"].fillna("")
        frames[s] = d

    # ========== Excel ==========
    xlsx = f"{OUT_DIR}/小波段策略_板块评估.xlsx"
    writer = pd.ExcelWriter(xlsx, engine="openpyxl")
    cols = ["date", "close", "板块风险", "大盘风险", "抄底狂热", "追涨狂热", "恐慌值",
            "买入预估成功率_5d", "disp_ma250", "vol_ratio", "signal", "reason"]

    for s in SECTORS:
        d = frames[s].copy()
        d["date"] = d["date"].astype(str).str[:10]
        d["disp_ma250"] = (d["disp_ma250"] * 100).round(1)
        d["vol_ratio"] = d["vol_ratio"].round(2)
        out = d[cols].rename(columns={
            "date": "日期", "close": "收盘", "disp_ma250": "偏离年线%",
            "vol_ratio": "量比", "买入预估成功率_5d": "买入预估成功率_5d"})
        out.to_excel(writer, sheet_name=s, index=False, startrow=1)

    # 买卖点明细
    detail = []
    for s in SECTORS:
        d = frames[s]; tr, _, _, _ = backtest(d, params)
        for t in tr:
            er = d.iloc[t["entry_i"]]; xr = d.iloc[t["exit_i"]]
            detail.append({
                "板块": s, "买入日期": t["entry_date"], "买入类型": er["signal"],
                "买入收盘": round(t["entry_close"], 1),
                "买入·板块风险": er["板块风险"], "买入·大盘风险": er["大盘风险"],
                "买入·恐慌值": er["恐慌值"], "买入·抄底狂热": er["抄底狂热"],
                "买入·预估成功率_5d": er["买入预估成功率_5d"],
                "买入理由": reason(er, er["signal"]),
                "卖出日期": t["exit_date"], "退出原因": t["exit_reason"],
                "卖出收盘": round(t["exit_close"], 1),
                "持仓(日)": t["hold_days"], "收益%": round(t["pnl"] * 100, 1)})
    pd.DataFrame(detail).to_excel(writer, sheet_name="买卖点明细", index=False)

    # 策略参数与回测
    param_rows = [
        ("长线买入·年线折价 lt_disp", params["lt_disp"]),
        ("长线买入·预估成功率阈值 lt_prob", params["lt_prob"]),
        ("买入A·局部恐慌阈值 buy_panic_l", params["buy_panic_l"]),
        ("买入A·局部板块风险上限 buy_risk_l", params["buy_risk_l"]),
        ("买入·大盘风险上限 buy_mkt", params["buy_mkt"]),
        ("买入·预估成功率阈值 buy_prob", params["buy_prob"]),
        ("买入B·全局板块风险上限 buy_risk_g", params["buy_risk_g"]),
        ("买入B·超卖阈值 oversold", params["oversold"]),
        ("卖出·追涨狂热阈值 sell_chase", params["sell_chase"]),
        ("卖出·板块风险阈值 sell_risk", params["sell_risk"]),
        ("硬止损 stop", params["stop"]), ("移动止盈回撤 trail", params["trail"]),
        ("同方向最小间隔 space", params["space"]), ("时间止损 tstop", params["tstop"]),
        ("迭代1综合评分", bp["iter1"]["score"]), ("迭代2综合评分", bp["iter2"]["score"]),
    ]
    pd.DataFrame(param_rows, columns=["参数", "取值"]).to_excel(
        writer, sheet_name="策略参数与回测", index=False, startrow=1)

    summ = pd.DataFrame([
        {"板块": s, "交易次数": stats[s]["trades"], "胜率%": stats[s]["win_rate"],
         "策略收益%": stats[s]["total_return"], "年化%": stats[s]["annual_return"],
         "最大回撤%": stats[s]["max_dd"], "持有不动%": stats[s]["buyhold_return"],
         "超额收益%": stats[s]["excess"], "平均持仓(日)": stats[s]["avg_hold"],
         "5日准确率%": stats[s]["buy_acc5"], "7日准确率%": stats[s]["buy_acc7"],
         "14日准确率%": stats[s]["buy_acc14"],
         "买点(含长线)": stats[s]["n_buy"], "卖点": stats[s]["n_sell"]}
        for s in SECTORS])
    summ = summ.replace({np.nan: "—"})
    summ.to_excel(writer, sheet_name="策略参数与回测", index=False, startrow=len(param_rows) + 4)
    writer.close()

    # ========== 样式 ==========
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    buy_fill = PatternFill("solid", fgColor="C6EFCE")
    sell_fill = PatternFill("solid", fgColor="FFC7CE")
    lt_fill  = PatternFill("solid", fgColor="FFEB9C")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for s in SECTORS:
        ws = wb[s]
        ws.cell(1, 1, f"【{s}】小波段策略 · 六大字段与买卖点信号").font = Font(bold=True, size=11)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(2, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        sig_col = cols.index("signal") + 1
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, sig_col).value
            if v in ("买入", "长线买入", "卖出"):
                fill = lt_fill if v == "长线买入" else (buy_fill if v == "买入" else sell_fill)
                for c in range(1, len(cols) + 1): ws.cell(r, c).fill = fill
        ws.freeze_panes = "A3"
        widths = [11, 9, 9, 9, 9, 9, 9, 14, 10, 8, 10, 70]
        for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb["买卖点明细"]
    for c in range(1, 17):
        cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 3).value
        fill = lt_fill if v == "长线买入" else (buy_fill if v == "买入" else None)
        if fill:
            for c in range(1, 17): ws.cell(r, c).fill = fill
    ws.freeze_panes = "A2"

    ws = wb["策略参数与回测"]
    ws.cell(1, 1, "小波段策略最终参数（两轮网格迭代优化）").font = Font(bold=True, size=11)
    for r in range(2, 2 + len(param_rows)):
        for c in (1, 2):
            cell = ws.cell(r, c)
            cell.fill = hdr_fill
            cell.font = hdr_font if c == 1 else Font(bold=True)
            cell.border = border
    sh = len(param_rows) + 4
    ws.cell(sh, 1, "各板块回测表现汇总").font = Font(bold=True, size=11)
    hr = sh + 1
    for c in range(1, 15):
        cell = ws.cell(hr, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in range(hr + 1, ws.max_row + 1):
        for c in range(1, 15): ws.cell(r, c).border = border
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 14
    for i in range(3, 15): ws.column_dimensions[get_column_letter(i)].width = 12
    wb.save(xlsx)
    print(f"[OK] Excel: {xlsx}")

    # ========== MD ==========
    md = []
    A = md.append
    A("# 小波段策略 · 每日运行报告")
    A("")
    A(f"> 生成日期：{pd.Timestamp.today().strftime('%Y-%m-%d')}")
    A(f"> 评估板块：半导体 / 消费电子 / 通信 / 电池 / 电力 / 煤炭（同花顺行业指数）")
    A("")
    A("## 信号规则（最终参数）")
    A("")
    A("| 信号 | 触发条件 |")
    A("|------|------|")
    A(f"| 买入A（恐慌企稳） | 局部恐慌≥{params['buy_panic_l']}，缩量企稳，局部风险<{params['buy_risk_l']}，大盘<{params['buy_mkt']}，预估≥{params['buy_prob']}% |")
    A(f"| 买入B（超卖缩量） | 超卖≥{params['oversold']}，缩量，全局风险<{params['buy_risk_g']}，大盘<{params['buy_mkt']}，预估≥{params['buy_prob']}% |")
    A(f"| 卖出 | 追涨狂热≥{params['sell_chase']}，板块风险≥{params['sell_risk']} |")
    A(f"| 风控 | 硬止损{params['stop']}，移动止盈回撤{params['trail']}(盈利1%启用)，间隔{params['space']}日，时间止损{params['tstop']}日 |")
    A("")
    A("## 各板块回测")
    A("")
    A("| 板块 | 交易数 | 胜率% | 策略收益% | 最大回撤% | 超额% | 持仓(日) | 5日准确率% | 7日准确率% |")
    A("|------|------|------|------|------|------|------|------|------|")
    for s in SECTORS:
        v = stats[s]
        A(f"| {s} | {v['trades']} | {v['win_rate']} | {v['total_return']} | {v['max_dd']} | {v['excess']} | {v['avg_hold']} | {nv(v['buy_acc5'])} | {nv(v['buy_acc7'])} |")
    A(f"| **均值** | **{avg(stats,'trades'):.0f}** | **{avg(stats,'win_rate')}** | **{avg(stats,'total_return')}** | **{avg(stats,'max_dd')}** | **{avg(stats,'excess')}** | **{avg(stats,'avg_hold')}** | **{avg(stats,'buy_acc5')}** | **{avg(stats,'buy_acc7')}** |")
    A("")

    det = pd.DataFrame(detail)
    ec = det["退出原因"].value_counts().to_dict()
    A(f"**退出原因（{len(det)}笔）**: 移动止盈={ec.get('移动止盈',0)} | 硬止损={ec.get('硬止损',0)} | 卖出信号={ec.get('卖出信号',0)} | 时间止损={ec.get('时间止损',0)}")
    A("")
    A("## 典型买卖点")
    A("")
    A("| 板块 | 买入日期 | 类型 | 卖出日期 | 退出原因 | 收益% | 预估5日成功率% |")
    A("|------|------|------|------|------|------|------|")
    for s in SECTORS:
        sub = det[det["板块"] == s]
        good = sub[(sub["买入·预估成功率_5d"] > 50) & (sub["收益%"] > 0)]
        row = good.sort_values("收益%", ascending=False).iloc[0] if len(good) else sub.sort_values("收益%", ascending=False).iloc[0]
        A(f"| {s} | {row['买入日期']} | {row['买入类型']} | {row['卖出日期']} | {row['退出原因']} | {row['收益%']} | {row['买入·预估成功率_5d']} |")
    A("")
    A("> 完整数据见配套 Excel。")
    A("")
    A("## 迭代评分")
    A(f"- 迭代1: {bp['iter1']['score']}")
    A(f"- 迭代2: {bp['iter2']['score']}")
    A("")
    A("---")
    A(f"*自动生成于 {pd.Timestamp.today().strftime('%Y-%m-%d %H:%M')}*")

    mdf = f"{OUT_DIR}/小波段_策略参数.md"
    with open(mdf, "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    print(f"[OK] MD: {mdf}")

if __name__ == "__main__":
    build()
