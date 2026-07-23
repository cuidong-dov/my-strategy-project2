#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
A股板块评估系统 · 小波段v2输出生成器
读取v2策略引擎输出的 signal CSV + stats/params JSON，生成 Excel 和 MD。
"""
import pandas as pd, numpy as np, json, os, sys, statistics as st, importlib.util

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
OUT_DIR  = os.path.join(BASE_DIR, "output")
SECTORS  = ["半导体", "消费电子", "通信", "电池", "电力", "煤炭"]

# 导入v2策略引擎
_engine = os.path.join(BASE_DIR, "scripts", "02_swing_engine_v2.py")
_spec = importlib.util.spec_from_file_location("engine_v2", _engine)
_mod = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(_mod)
backtest, reason = _mod.backtest, _mod.reason

def nv(v):
    return "—" if v is None else v

def avg(dd, m):
    vals = [dd[s][m] for s in SECTORS if dd[s][m] is not None]
    return round(st.mean(vals), 1) if vals else "—"

def build():
    stats_d = json.load(open(f"{OUT_DIR}/swing_v2_stats.json", encoding="utf-8"))
    bp       = json.load(open(f"{OUT_DIR}/swing_v2_params.json", encoding="utf-8"))
    params = bp["params"]

    # 加载信号CSV
    frames = {}
    for s in SECTORS:
        d = pd.read_csv(f"{OUT_DIR}/swing_v2_{s}_signal.csv", parse_dates=["date"])
        d["signal"] = d["signal"].fillna("")
        frames[s] = d

    # ========== Excel ==========
    xlsx = f"{OUT_DIR}/小波段策略v2_板块评估.xlsx"
    writer = pd.ExcelWriter(xlsx, engine="openpyxl")
    cols = ["date", "close", "板块风险", "大盘风险", "抄底狂热", "追涨狂热", "恐慌值",
            "买入预估成功率_5d", "buy_score", "disp_ma250", "vol_ratio", "vol_quality",
            "pullback_from_high", "macd_bull_div", "vol_bull_div",
            "macd_bear_div", "vol_bear_div", "signal", "reason"]

    for s in SECTORS:
        d = frames[s].copy()
        d["date"] = d["date"].astype(str).str[:10]
        d["disp_ma250"] = (d["disp_ma250"] * 100).round(1)
        d["vol_ratio"] = d["vol_ratio"].round(2)
        d["vol_quality"] = d["vol_quality"].round(2)
        d["pullback_from_high"] = (d["pullback_from_high"] * 100).round(1)
        available = [c for c in cols if c in d.columns]
        out = d[available].rename(columns={
            "date": "日期", "close": "收盘", "buy_score": "买点评分",
            "disp_ma250": "偏离年线%",
            "vol_ratio": "量比", "买入预估成功率_5d": "预估成功率_5d",
            "vol_quality": "量价质量", "pullback_from_high": "回调幅度%",
            "macd_bull_div": "MACD底背离", "vol_bull_div": "量价底背离",
            "macd_bear_div": "MACD顶背离", "vol_bear_div": "量价顶背离"})
        out.to_excel(writer, sheet_name=s, index=False, startrow=1)

    # 买卖点明细
    detail = []
    for s in SECTORS:
        d = frames[s]; tr, _, _, _ = backtest(d, params)
        for t in tr:
            er = d.iloc[t["entry_i"]]; xr = d.iloc[t["exit_i"]]
            detail.append({
                "板块": s, "买入日期": t["entry_date"], "买入类型": er["signal"],
                "买点评分": er.get("buy_score", "—"),
                "买入收盘": round(t["entry_close"], 1),
                "买入·板块风险": er["板块风险"], "买入·大盘风险": er["大盘风险"],
                "买入·恐慌值": er["恐慌值"], "买入·抄底狂热": er["抄底狂热"],
                "买入·预估成功率_5d": er["买入预估成功率_5d"],
                "买入·MACD底背离": er.get("macd_bull_div", 0),
                "买入·量价底背离": er.get("vol_bull_div", 0),
                "买入理由": reason(er, er["signal"]),
                "卖出日期": t["exit_date"], "退出原因": t["exit_reason"],
                "卖出收盘": round(t["exit_close"], 1),
                "持仓(日)": t["hold_days"], "收益%": round(t["pnl"] * 100, 1)})
    pd.DataFrame(detail).to_excel(writer, sheet_name="买卖点明细", index=False)

    # 策略参数与回测
    param_rows = [
        ("=== v1保留参数 ===", ""),
        ("长线买入·年线折价 lt_disp", params["lt_disp"]),
        ("长线买入·预估成功率阈值 lt_prob", params["lt_prob"]),
        ("买入A·局部恐慌阈值 buy_panic_l", params["buy_panic_l"]),
        ("买入A·局部板块风险上限 buy_risk_l", params["buy_risk_l"]),
        ("买入·大盘风险上限 buy_mkt", params["buy_mkt"]),
        ("买入·预估成功率阈值 buy_prob", params["buy_prob"]),
        ("买入B·全局板块风险上限 buy_risk_g", params["buy_risk_g"]),
        ("卖出·追涨狂热阈值 sell_chase", params["sell_chase"]),
        ("卖出·板块风险阈值 sell_risk", params["sell_risk"]),
        ("买入B·超卖阈值 oversold", params["oversold"]),
        ("硬止损 stop", params["stop"]), ("移动止盈回撤 trail", params["trail"]),
        ("同方向最小间隔 space", params["space"]), ("时间止损 tstop", params["tstop"]),
        ("", ""),
        ("=== v2新增参数 ===", ""),
        ("量价质量下限 vol_quality_min", params["vol_quality_min"]),
        ("回调幅度上限 pullback_max", params["pullback_max"]),
        ("放量下跌占比上限 heavy_down_max", params["heavy_down_max"]),
        ("底背离·连续缩量天数 shrink_streak", params["shrink_streak"]),
        ("", ""),
        ("=== 迭代评分 ===", ""),
        ("v2迭代1综合评分", bp["iter1"]["score"]),
        ("v2迭代2综合评分", bp["iter2"]["score"]),
    ]
    pd.DataFrame(param_rows, columns=["参数", "取值"]).to_excel(
        writer, sheet_name="策略参数与回测", index=False, startrow=1)

    summ = pd.DataFrame([
        {"板块": s, "交易次数": stats_d[s]["trades"], "胜率%": stats_d[s]["win_rate"],
         "策略收益%": stats_d[s]["total_return"], "年化%": stats_d[s]["annual_return"],
         "最大回撤%": stats_d[s]["max_dd"], "持有不动%": stats_d[s]["buyhold_return"],
         "超额收益%": stats_d[s]["excess"], "平均持仓(日)": stats_d[s]["avg_hold"],
         "5日准确率%": stats_d[s]["buy_acc5"], "7日准确率%": stats_d[s]["buy_acc7"],
         "14日准确率%": stats_d[s]["buy_acc14"],
         "买点(含长线)": stats_d[s]["n_buy"], "卖点": stats_d[s]["n_sell"]}
        for s in SECTORS])
    summ = summ.replace({np.nan: "—"})
    summ.to_excel(writer, sheet_name="策略参数与回测", index=False, startrow=len(param_rows) + 4)
    writer.close()

    # ========== 样�� ==========
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(xlsx)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(color="FFFFFF", bold=True)
    buy_fill = PatternFill("solid", fgColor="C6EFCE")
    sell_fill = PatternFill("solid", fgColor="FFC7CE")
    lt_fill  = PatternFill("solid", fgColor="FFEB9C")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for s in SECTORS:
        ws = wb[s]
        ws.cell(1, 1, f"【{s}】小波段v2策略 · 含买点评分+量价背离").font = Font(bold=True, size=11)
        sheet_cols = [c for c in cols if c in frames[s].columns]
        for c_idx in range(1, len(sheet_cols) + 1):
            cell = ws.cell(2, c_idx); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        sig_col = sheet_cols.index("signal") + 1 if "signal" in sheet_cols else len(sheet_cols)
        score_col = sheet_cols.index("buy_score") + 1 if "buy_score" in sheet_cols else None
        high_fill = PatternFill("solid", fgColor="00B050")  # >=7分 深绿
        mid_fill  = PatternFill("solid", fgColor="92D050")  # 5-6分 中绿
        low_fill  = PatternFill("solid", fgColor="C6EFCE")  # <=4分 浅绿
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, sig_col).value
            if v in ("买入", "长线买入", "卖出"):
                fill = lt_fill if v == "长线买入" else (buy_fill if v == "买入" else sell_fill)
                for c in range(1, len(sheet_cols) + 1): ws.cell(r, c).fill = fill
            # 评分列颜色覆盖（买入行）
            if v in ("买入", "长线买入") and score_col:
                sv = ws.cell(r, score_col).value
                if sv is not None:
                    try:
                        sv = float(sv)
                        if sv >= 7: ws.cell(r, score_col).fill = high_fill
                        elif sv >= 5: ws.cell(r, score_col).fill = mid_fill
                        else: ws.cell(r, score_col).fill = low_fill
                    except: pass
        ws.freeze_panes = "A3"
        widths = [11, 9, 9, 9, 9, 9, 9, 13, 7, 10, 8, 8, 9, 9, 9, 9, 9, 10, 80]
        for i, w in enumerate(widths[:len(sheet_cols)], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb["买卖点明细"]
    ncol = len(detail[0]) if detail else 17
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 3).value
        fill = lt_fill if v == "长线买入" else (buy_fill if v == "买入" else None)
        if fill:
            for c in range(1, ncol + 1): ws.cell(r, c).fill = fill
    ws.freeze_panes = "A2"

    ws = wb["策略参数与回测"]
    ws.cell(1, 1, "小波段v2策略最终参数（两轮网格迭代优化）").font = Font(bold=True, size=11)
    for r in range(2, 2 + len(param_rows)):
        for c in (1, 2):
            cell = ws.cell(r, c)
            if ws.cell(r, 1).value and str(ws.cell(r, 1).value).startswith("==="):
                cell.font = Font(bold=True, color="1F4E78")
            else:
                cell.fill = hdr_fill
                cell.font = hdr_font if c == 1 else Font(bold=True)
            cell.border = border
    sh = len(param_rows) + 4
    ws.cell(sh, 1, "各板块回测表现汇总").font = Font(bold=True, size=11)
    hr = sh + 1
    for c in range(1, 15):
        cell = ws.cell(hr, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in range(hr + 1, ws.max_row + 1):
        for c in range(1, 15): ws.cell(r, c).border = border
    ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 14
    for i in range(3, 15): ws.column_dimensions[get_column_letter(i)].width = 12
    wb.save(xlsx)
    print(f"[OK] Excel: {xlsx}")

    # ========== MD ==========
    md = []
    A = md.append
    A("# 小波段策略 v2 · 完整参数文档")
    A("")
    A(f"> 生成日期：{pd.Timestamp.today().strftime('%Y-%m-%d')}")
    A("> 版本：v2（新增量价分析、回调到位、MACD背离）")
    A(f"> 最终评分：{bp['score']}")
    A("")
    A("## 一、v2 vs v1 核心变化")
    A("")
    A("| 维度 | v1 | v2 |")
    A("|------|-----|-----|")
    A("| 买入信号 | 2种（恐慌企稳A、超卖缩量B） | 3种（+底背离买入C） |")
    A("| 卖出信号 | 追涨狂热+板块风险 | +顶背离辅助卖出 |")
    A("| 量价分析 | 仅量比+缩量判断 | 新增量价质量、放量下跌识别 |")
    A("| 回调分析 | 仅5日/20日回撤 | 新增斐波那契回调位、前低支撑 |")
    A("| 背离检测 | 无 | MACD底/顶背离、量价底/顶背离 |")
    A("| 成功率因子 | 5因子 | 8因子（+量价质量、回调深度、MACD趋势） |")
    A("")
    A("## 二、覆盖标的")
    A("")
    A("| 板块名称 | THS行业名 | THS代码 |")
    A("|----------|----------|---------|")
    A("| 半导体 | 半导体 | 881121 |")
    A("| 消费电子 | 消费电子 | 881124 |")
    A("| 通信 | 通信设备 | 881129 |")
    A("| 电池 | 电池 | 881281 |")
    A("| 电力 | 电力 | 881145 |")
    A("| 煤炭 | 煤炭开采加工 | 881105 |")
    A("")
    A("## 三、v2新增指标定义")
    A("")
    A("### 3.1 量价分析")
    A("")
    A("| 指标 | 计算方式 | 含义 |")
    A("|------|---------|------|")
    A("| `vol_quality` | (涨放量-跌缩量)均值/总均值 | 量价配合度，正值=健康 |")
    A("| `vol_heavy_down` | 下跌日且量比>1.3 | 放量下跌=恐慌出逃 |")
    A("| `heavy_down_5d` | 近5日放量下跌天数占比 | 越高越危险 |")
    A("| `vol_shrink_streak` | 连续缩量天数 | 缩量越久=抛压越枯竭 |")
    A("")
    A("### 3.2 回调到位")
    A("")
    A("| 指标 | 计算方式 | 含义 |")
    A("|------|---------|------|")
    A("| `pullback_from_high` | close/20日最高-1 | 从近期高点回调幅度 |")
    A("| `near_fib_618/500/382` | 是否接近斐波那契支撑位 | 技术支撑参考 |")
    A("| `near_prev_low` | 是否接近20日前低 | 二次探底信号 |")
    A("| `pullback_deep` | 回调幅度>1.5倍ATR | 回调是否充分 |")
    A("")
    A("### 3.3 MACD背离")
    A("")
    A("| 指标 | 含义 | 信号 |")
    A("|------|------|------|")
    A("| `macd_bull_div` | MACD底背离：价新低，DIF没新低 | 下跌动能衰竭，看涨 |")
    A("| `macd_bear_div` | MACD顶背离：价新高，DIF没新高 | 上涨动能衰竭，看跌 |")
    A("| `vol_bull_div` | 量价底背离：价新低，量更小 | 抛压枯竭，看涨 |")
    A("| `vol_bear_div` | 量价顶背离：价新高，量更小 | 追涨无力，看跌 |")
    A("")
    A("## 四、信号规则（v2最优参数）")
    A("")
    A("### 4.1 完整参数表")
    A("")
    A("| 参数 | v1值 | v2值 | 变化 | 含义 |")
    A("|------|------|------|------|------|")
    A(f"| buy_panic_l | 58 | {params['buy_panic_l']} | ↓ | 买入A恐慌阈值放宽 |")
    A(f"| buy_prob | 43 | {params['buy_prob']} | ↓ | 预估成功率阈值放宽 |")
    A(f"| buy_risk_g | 50 | {params['buy_risk_g']} | ↓ | 买入B风险更严格 |")
    A(f"| sell_chase | 74 | {params['sell_chase']} | ↑ | 卖出追涨阈值提高 |")
    A(f"| sell_risk | 74 | {params['sell_risk']} | ↑ | 卖出风险阈值提高 |")
    A(f"| oversold | 0.50 | {params['oversold']} | ↓ | 超卖阈值放宽 |")
    A(f"| vol_quality_min | — | {params['vol_quality_min']} | 新增 | 量价质量下限 |")
    A(f"| pullback_max | — | {params['pullback_max']} | 新增 | 回调幅度上限 |")
    A(f"| heavy_down_max | — | {params['heavy_down_max']} | 新增 | 放量下跌上限 |")
    A(f"| shrink_streak | — | {params['shrink_streak']} | 新增 | 底背离需连续缩量天数 |")
    A("")
    A("### 4.2 买入信号")
    A("")
    A("**买入A（恐慌企稳+v2增强）**：")
    A(f"- 局部恐慌 ≥ {params['buy_panic_l']}，缩量企稳")
    A(f"- 局部风险 < {params['buy_risk_l']}，大盘 < {params['buy_mkt']}")
    A(f"- 预估成功率 ≥ {params['buy_prob']}%")
    A(f"- **v2新增**：量价质量 > {params['vol_quality_min']}，回调幅度 < {params['pullback_max']}，近5日放量跌占比 < {params['heavy_down_max']}")
    A("")
    A("**买入B（超卖缩量+v2增强）**：")
    A(f"- 超卖 ≥ {params['oversold']}，缩量")
    A(f"- 全局风险 < {params['buy_risk_g']}，大盘 < {params['buy_mkt']}")
    A(f"- 预估成功率 ≥ {params['buy_prob']}%")
    A(f"- **v2新增**：量价质量 > {params['vol_quality_min']}，回调幅度 < {params['pullback_max']}")
    A("")
    A("**买入C（v2新增·底背离买入）**：")
    A("- MACD底背离 或 量价底背离")
    A(f"- 全局风险 < {params['buy_risk_g']+10}，大盘 < {params['buy_mkt']}")
    A(f"- 预估成功率 ≥ {params['buy_prob']-5}%")
    A(f"- 连续缩量 ≥ {params['shrink_streak']}天")
    A("")
    A("### 4.3 卖出信号（v2增强）")
    A("")
    A(f"- 基础：追涨狂热 ≥ {params['sell_chase']}，板块风险 ≥ {params['sell_risk']}")
    A(f"- **v2新增**：MACD顶背离 或 量价顶背离，且风险 ≥ {params['sell_risk']-5}")
    A("")
    A("## 五、回测表现（2021-01 ~ 2026-07-22）")
    A("")
    A("| 板块 | 交易数 | 胜率% | 策略收益% | 最大回撤% | 超额% | 持仓(日) | 5日准确% | 7日准确% |")
    A("|------|--------|-------|-----------|-----------|-------|----------|----------|----------|")
    for s in SECTORS:
        v = stats_d[s]
        A(f"| {s} | {v['trades']} | {v['win_rate']} | {v['total_return']} | {v['max_dd']} | {v['excess']} | {v['avg_hold']} | {nv(v['buy_acc5'])} | {nv(v['buy_acc7'])} |")
    A(f"| **均值** | **{avg(stats_d,'trades'):.0f}** | **{avg(stats_d,'win_rate')}** | **{avg(stats_d,'total_return')}** | **{avg(stats_d,'max_dd')}** | **{avg(stats_d,'excess')}** | **{avg(stats_d,'avg_hold')}** | **{avg(stats_d,'buy_acc5')}** | **{avg(stats_d,'buy_acc7')}** |")
    A("")

    det = pd.DataFrame(detail)
    ec = det["退出原因"].value_counts().to_dict()
    A(f"**退出原因（{len(det)}笔）**: 移动止盈={ec.get('移动止盈',0)} | 硬止损={ec.get('硬止损',0)} | 卖出信号={ec.get('卖出信号',0)} | 时间止损={ec.get('时间止损',0)}")
    A("")
    A("## 六、v1 vs v2 对比")
    A("")
    A("| 指标 | v1 | v2 | 变化 |")
    A("|------|-----|-----|------|")
    # 加载v1数据对比
    v1_stats = json.load(open(f"{OUT_DIR}/swing_stats.json", encoding="utf-8"))
    v1_trades = sum(v1_stats[s]["trades"] for s in SECTORS)
    v2_trades = sum(stats_d[s]["trades"] for s in SECTORS)
    v1_wr = avg(v1_stats, "win_rate")
    v2_wr = avg(stats_d, "win_rate")
    v1_ret = avg(v1_stats, "total_return")
    v2_ret = avg(stats_d, "total_return")
    v1_mdd = avg(v1_stats, "max_dd")
    v2_mdd = avg(stats_d, "max_dd")
    v1_a7 = avg(v1_stats, "buy_acc7")
    v2_a7 = avg(stats_d, "buy_acc7")
    A(f"| 总交易数 | {v1_trades} | {v2_trades} | {'+' if v2_trades > v1_trades else ''}{v2_trades - v1_trades} |")
    A(f"| 胜率 | {v1_wr}% | {v2_wr}% | {'+' if v2_wr > v1_wr else ''}{round(v2_wr - v1_wr, 1)}% |")
    A(f"| 平均收益 | {v1_ret}% | {v2_ret}% | {'+' if v2_ret > v1_ret else ''}{round(v2_ret - v1_ret, 1)}% |")
    A(f"| 平均回撤 | {v1_mdd}% | {v2_mdd}% | {'改善' if v2_mdd > v1_mdd else '变差'}{round(abs(v2_mdd) - abs(v1_mdd), 1)}% |")
    A(f"| 7日准确率 | {v1_a7}% | {v2_a7}% | {'+' if v2_a7 > v1_a7 else ''}{round(v2_a7 - v1_a7, 1)}% |")
    A(f"| 综合评分 | 0.8495 | {bp['score']} | +{round(bp['score'] - 0.8495, 4)} |")
    A("")
    A("## 七、换对话复现指南")
    A("")
    A("```bash")
    A("cd /workspace/a-share-sector-eval")
    A("python3.11 scripts/01_fetch_data.py")
    A("python3.11 scripts/02_swing_engine_v2.py")
    A("python3.11 scripts/03_build_output_v2.py")
    A("```")
    A("")
    A("---")
    A(f"*自动生成于 {pd.Timestamp.today().strftime('%Y-%m-%d %H:%M')}*")

    mdf = f"{OUT_DIR}/小波段v2_策略参数.md"
    with open(mdf, "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    print(f"[OK] MD: {mdf}")

if __name__ == "__main__":
    build()
